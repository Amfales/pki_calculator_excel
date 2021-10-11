import openpyxl
import typing

from plugins.base_plugin import AbstractBasePlugin
from plugins.data_manager import DataManager
from pki_manager.pki import LiteralPKI

class XLSXPluginFactory():
  def __init__(self, num_years: int, row_base: int, col_base: int):
    self.num_years = num_years
    self.row_base = row_base
    self.col_base = col_base

  
  def create_plugin(self, sheetname: str, data_name: str, base_row: int):
    outer_self = self
    class XLSXPlugin(AbstractBasePlugin):
      @property
      def datapoint_name(self) -> str:
          return data_name

      @property
      def dependencies(self) -> list[str]:
          return []

      def do_work(self, workbook: openpyxl.Workbook, data: dict[str, typing.Any]) -> typing.Any:
        income_idx = workbook.sheetnames.index(sheetname)
        income_sheet = workbook.worksheets[income_idx] if income_idx >= 0 else None
        vals = {}
        for i in range(outer_self.num_years):
          date_cell = income_sheet.cell(outer_self.row_base, outer_self.col_base + i)
          val_cell = income_sheet.cell(base_row, outer_self.col_base + i)
          vals[date_cell.value] = LiteralPKI(val_cell.value) if val_cell.value is not None else LiteralPKI(0)
        return vals
    return XLSXPlugin()


class IncomePluginFactory():
  def __init__(self, num_years=1):
    self.num_years = num_years

  def create_plugin(self, data_name: str, dependencies: list[str], base_row: int):
    class IncomePlugin(AbstractBasePlugin):
      def __init__(self, num_years=1):
        self.num_years = num_years

      @property
      def datapoint_name(self) -> str:
        return data_name

      @property
      def dependencies(self) -> list[str]:
          return dependencies

      def do_work(self, workbook: openpyxl.Workbook, data: dict[str, typing.Any]) -> typing.Any:
        income_idx = workbook.sheetnames.index('INCOME STATEMENTS')
        income_sheet = workbook.worksheets[income_idx] if income_idx >= 0 else None
        vals = {}
        for i in range(self.num_years):
          date_cell = income_sheet.cell(3, 2 + i)
          val_cell = income_sheet.cell(base_row, 2 + i)
          vals[date_cell.value] = LiteralPKI(val_cell.value) if val_cell.value is not None else LiteralPKI(0)
        return vals
    return IncomePlugin(self.num_years)